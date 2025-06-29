import os
import openpyxl

directorio_pdfs = "C:\\Users\\Pedro\\Desktop\\LotoIA\\LOTERIA NACIONAL\\Historico-Resultados-LoteriaNacional"
anios = os.listdir(directorio_pdfs)

# Crear un nuevo libro de Excel
libro_excel = openpyxl.Workbook()
hoja = libro_excel.active

# Escribir los encabezados en la hoja de Excel
hoja['A1'] = 'Año'
hoja['B1'] = 'Número'
hoja['C1'] = 'Cantidad'
hoja['D1'] = 'Fecha'

fila = 2

for anio in anios:
    if anio.isdigit() and int(anio) >= 1812:
        directorio_anio = os.path.join(directorio_pdfs, anio)
        archivos_pdf = os.listdir(directorio_anio)
        
        for archivo_pdf in archivos_pdf:
            ruta_archivo = os.path.join(directorio_anio, archivo_pdf)
            
            try:
                # Abrir el archivo PDF y extraer los datos
                with open(ruta_archivo, "r", encoding="latin-1") as file:
                    texto = file.read()
                    
                    # Procesar el texto y extraer los datos relevantes
                    lineas = texto.split('\n')
                    for linea in lineas:
                        if linea.strip():
                            datos = linea.split(' ')
                            numero = datos[0].strip()
                            cantidad = ' '.join(datos[1:]).strip()
                            fecha = archivo_pdf.split('.')[2]
                            
                            # Escribir los datos en la hoja de Excel
                            hoja.cell(row=fila, column=1, value=anio)
                            hoja.cell(row=fila, column=2, value=numero)
                            hoja.cell(row=fila, column=3, value=cantidad)
                            hoja.cell(row=fila, column=4, value=fecha)
                            fila += 1
                
                print(f"Procesado: {ruta_archivo}")
            except Exception as e:
                print(f"Error al procesar: {ruta_archivo} - {str(e)}")

# Guardar el libro de Excel
libro_excel.save("datos_loteria.xlsx")