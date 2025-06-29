import os
import PyPDF2
import openpyxl

directorio_pdfs = "C:\\Users\\Pedro\\Desktop\\LotoIA\\LOTERÍA NACIONAL\\Histórico-Resultados-LoteriaNacional"
num_archivos = 1817

# Crear un nuevo libro de Excel
libro_excel = openpyxl.Workbook()
hoja = libro_excel.active

# Escribir los encabezados en la hoja de Excel
hoja['A1'] = 'Número'
hoja['B1'] = 'Cantidad'
hoja['C1'] = 'Fecha'

fila = 2

for i in range(1, num_archivos + 1):
    nombre_archivo = f"SM_LISTAOFICIAL.A2024.S0{i:02d}.pdf"
    ruta_archivo = os.path.join(directorio_pdfs, nombre_archivo)
    
    if os.path.exists(ruta_archivo):
        # Abrir el archivo PDF y extraer los datos
        with open(ruta_archivo, "rb") as file:
            lector_pdf = PyPDF2.PdfReader(file)
            pagina = lector_pdf.pages[0]
            texto = pagina.extract_text()
            
            # Procesar el texto y extraer los datos relevantes
            lineas = texto.split('\n')
            for linea in lineas:
                if ':' in linea:
                    datos = linea.split(':')
                    numero = datos[0].strip()
                    cantidad = datos[1].strip()
                    fecha = f"S0{i:02d}"
                    
                    # Escribir los datos en la hoja de Excel
                    hoja.cell(row=fila, column=1, value=numero)
                    hoja.cell(row=fila, column=2, value=cantidad)
                    hoja.cell(row=fila, column=3, value=fecha)
                    fila += 1
        
        print(f"Procesado: {ruta_archivo}")
    else:
        print(f"Archivo no encontrado: {ruta_archivo}")

# Guardar el libro de Excel
libro_excel.save("datos_loteria.xlsx")